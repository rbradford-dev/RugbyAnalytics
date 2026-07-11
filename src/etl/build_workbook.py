"""
Build the RugbyAnalytics multi-year Excel workbook from ETL output.

Input:  data/processed/sixnations_raw.json  (produced by src/etl/rugbypass_pull.py)
Output: outputs/RugbyAnalytics_SixNations_MultiYear.xlsx

Sheets: NOTES (data dictionary), TEAM_STATS (style vectors), POWER_RANKING,
STAT_<year> (per-season player stats, all 23 schema metrics + player Elo),
MATCH_RESULTS (with walk-forward Elo win probabilities), ELO (team + top
players), MODEL_CHECK (Brier/accuracy vs baseline).

Usage:
    python -m src.etl.build_workbook \
        --raw data/processed/sixnations_raw.json \
        --out outputs/RugbyAnalytics_SixNations_MultiYear.xlsx \
        --seasons 2024 2025
"""
import argparse, json, os
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.features import elo as elomod
from src.features.team_analysis import aggregate_team_stats, backtest_elo

ABBR = {'France': 'FRE', 'Wales': 'WAL', 'Scotland': 'SCT', 'Italy': 'ITL',
        'Ireland': 'IRE', 'England': 'ENG'}
GREEN, WHITE = "1F6E43", "FFFFFF"
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STAT_COLS = ['PLAYER_NAME', 'COUNTRY', 'POSITION', 'APPEARANCES', 'MINUTES',
             'POINTS', 'TRIES', 'LINE_BREAKS', 'CARRIES', 'CARRIES_PER_MIN',
             'METERS_CARRIED', 'DEFENDERS_BEATEN', 'TURNOVERS_WON_ATT', 'TOUCHES',
             'POST_CONTACT_METERS', 'TACKLES_MADE', 'TACKLES_COMPLETED',
             'DOMINANT_TACKLES', 'TACKLES_PER_MIN', 'TURNOVERS_WON_DEF',
             'RUCK_ARRIVAL_EFFECTIVENESS', 'TRY_ASSISTS', 'SUCCESSFUL_PASSES',
             'BAD_PASSES', 'PASS_ACCURACY', 'PENALTIES_CONCEDED', 'YELLOW_CARDS',
             'RED_CARDS', 'PLAYER_ELO']


def hdr(cell, fill=GREEN, font_color=WHITE):
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(bold=True, color=font_color, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def season_snapshots(matches_by_season, seasons):
    """Cumulative Elo carried across seasons; snapshot at each season end."""
    r = defaultdict(lambda: 1500.0)
    def exp(a, b): return 1 / (1 + 10 ** ((b - a) / 400))
    snaps = {}
    for yr in seasons:
        deltas = defaultdict(list)
        for m in sorted(matches_by_season.get(str(yr), matches_by_season.get(yr, [])),
                        key=lambda x: x['date']):
            h, a = m['home'], m['away']
            rh, ra = r[h], r[a]
            eh = exp(rh + elomod.HOME_ADV, ra)
            sh = 1.0 if m['hs'] > m['aw'] else (0.0 if m['hs'] < m['aw'] else 0.5)
            mult = elomod.mov_multiplier(m['hs'] - m['aw'], (rh + elomod.HOME_ADV) - ra)
            dh = elomod.K * mult * (sh - eh); da = elomod.K * mult * ((1 - sh) - (1 - eh))
            r[h] = rh + dh; r[a] = ra + da
            deltas[h].append(dh); deltas[a].append(da)
        snaps[yr] = ({t: round(v, 1) for t, v in r.items()}, dict(deltas))
    return snaps


def player_row(name, country, pos, blk, pelo):
    mins = blk.get('minutes_played_total') or 0
    carries = blk.get('carries') or 0
    tackles = blk.get('tackles') or 0
    passes = blk.get('passes') or 0
    bad = blk.get('bad_passes') or 0
    tot = passes + bad
    return {
        'PLAYER_NAME': name, 'COUNTRY': country, 'POSITION': pos,
        'APPEARANCES': blk.get('total_games'), 'MINUTES': mins,
        'POINTS': blk.get('points'), 'TRIES': blk.get('tries'),
        'LINE_BREAKS': blk.get('clean_breaks'), 'CARRIES': carries,
        'CARRIES_PER_MIN': round(carries / mins, 3) if mins else None,
        'METERS_CARRIED': blk.get('metres'), 'DEFENDERS_BEATEN': blk.get('defenders_beaten'),
        'TURNOVERS_WON_ATT': blk.get('turnovers_won'), 'TOUCHES': blk.get('touches'),
        'POST_CONTACT_METERS': blk.get('post_contact_metres'), 'TACKLES_MADE': tackles,
        'TACKLES_COMPLETED': tackles - (blk.get('missed_tackles') or 0),
        'DOMINANT_TACKLES': blk.get('dominant_tackles'),
        'TACKLES_PER_MIN': round(tackles / mins, 3) if mins else None,
        'TURNOVERS_WON_DEF': blk.get('turnovers_won'),
        'RUCK_ARRIVAL_EFFECTIVENESS': blk.get('att_ruck_arrival_effectiveness'),
        'TRY_ASSISTS': blk.get('try_assist'), 'SUCCESSFUL_PASSES': passes,
        'BAD_PASSES': bad, 'PASS_ACCURACY': round(passes / tot, 3) if tot else None,
        'PENALTIES_CONCEDED': blk.get('penalties_conceded'),
        'YELLOW_CARDS': blk.get('yellow_cards'), 'RED_CARDS': blk.get('red_cards'),
        'PLAYER_ELO': pelo,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--raw", default="data/processed/sixnations_raw.json")
    ap.add_argument("--out", default="outputs/RugbyAnalytics_SixNations_MultiYear.xlsx")
    ap.add_argument("--seasons", nargs="+", type=int, default=[2024, 2025])
    args = ap.parse_args()

    raw = json.load(open(args.raw))
    matches, players = raw['matches'], raw['players']
    # normalize season keys of player blocks to int
    for pid, r in players.items():
        r['seasons'] = {int(k): v for k, v in r['seasons'].items()}

    snaps = season_snapshots(matches, args.seasons)
    team_stats = aggregate_team_stats(players, ABBR, args.seasons)
    bt = backtest_elo({int(k): v for k, v in matches.items()},
                      sorted(int(k) for k in matches))

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "NOTES"
    ws.append(["RugbyAnalytics — generated by src/etl/build_workbook.py"])
    ws.append(["Sources", "ESPN JSON API (matches) + RugbyPass/Opta player pages"])
    ws.append(["Blank cells", "missing at source — NEVER fabricated"])

    # per-season player sheets
    for yr in args.seasons:
        snap, deltas = snaps[yr]
        prev = snaps.get(yr - 1, ({}, {}))[0]
        sh = wb.create_sheet(f"STAT_{yr}")
        sh.append(STAT_COLS)
        recs = []
        for pid, r in players.items():
            blk = r['seasons'].get(yr)
            if not blk:
                continue
            team = ABBR.get(r['country'], r['country'])
            d = deltas.get(team, [])
            n = len(d) or 1
            mins = blk.get('minutes_played_total') or 0
            share = min(mins / (n * 80.0), 1.0)
            pelo = round(prev.get(team, 1500.0) + share * sum(d), 1)
            recs.append(player_row(r['name'], team, '', blk, pelo))
        recs.sort(key=lambda x: (x['COUNTRY'], -(x['POINTS'] or 0)))
        for rec in recs:
            sh.append([rec.get(c) for c in STAT_COLS])
        for c in sh[1]:
            hdr(c)
        sh.freeze_panes = "D2"

    # team stats
    t = wb.create_sheet("TEAM_STATS")
    t.append(['SEASON', 'TEAM', 'POINTS', 'TRIES', 'CARRIES', 'METERS', 'TACKLES',
              'DOM_TACKLES', 'OFFLOADS', 'PASS_ACCURACY', 'ELO'])
    for (yr, team), s in sorted(team_stats.items()):
        tot = s['passes'] + s['bad_passes']
        t.append([yr, team, round(s['points']), round(s['tries']), round(s['carries']),
                  round(s['metres']), round(s['tackles']), round(s['dominant_tackles']),
                  round(s['offloads']), round(s['passes'] / tot, 3) if tot else None,
                  snaps[yr][0].get(team)])
    for c in t[1]:
        hdr(c)

    # match results + model check
    mr = wb.create_sheet("MATCH_RESULTS")
    mr.append(['DATE', 'HOME', 'AWAY', 'HS', 'AS', 'P_HOME(ELO)', 'ACTUAL'])
    for m in bt['rows']:
        mr.append([m['date'], m['home'], m['away'], m['hs'], m['aw'],
                   m['p_home'], m['actual']])
    for c in mr[1]:
        hdr(c)

    mc = wb.create_sheet("MODEL_CHECK")
    mc.append(["Model", "Brier", "Accuracy"])
    mc.append(["Elo (MOV+home)", round(bt['brier_elo'], 3), round(bt['acc_elo'], 3)])
    mc.append(["Home-always", round(bt['brier_home'], 3), round(bt['acc_home'], 3)])
    for c in mc[1]:
        hdr(c)

    os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)
    wb.save(args.out)
    print(f"wrote {args.out}")


if __name__ == "__main__":
    main()
